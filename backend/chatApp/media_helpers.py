# chatApp/media_helpers.py
import os
import logging
import mimetypes
from django.conf import settings
from PIL import Image

logger = logging.getLogger(__name__)

# Define MediaFileType here to avoid circular imports
class MediaFileType:
    IMAGE = 'image'
    VIDEO = 'video'
    AUDIO = 'audio'
    VOICE_NOTE = 'voice_note'
    DOCUMENT = 'document'
    PDF = 'pdf'
    WORD = 'word'
    EXCEL = 'excel'
    OTHER = 'other'
    
    CHOICES = [
        (IMAGE, 'Image'),
        (VIDEO, 'Video'),
        (AUDIO, 'Audio'),
        (VOICE_NOTE, 'Voice Note'),
        (DOCUMENT, 'Document'),
        (PDF, 'PDF'),
        (WORD, 'Word Document'),
        (EXCEL, 'Excel Spreadsheet'),
        (OTHER, 'Other'),
    ]

import os
import io
import logging
import mimetypes
from django.conf import settings
from PIL import Image

logger = logging.getLogger(__name__)


def get_mime_type(file_obj):
    """Get mime type using mimetypes module"""
    mime_type, _ = mimetypes.guess_type(file_obj.name)
    return mime_type or 'application/octet-stream'


def determine_file_type(filename, mime_type=None):
    """
    Determine the media file type.
    Special case: browser-recorded voice notes come in as audio/webm or video/webm
    but their filename starts with 'voice-', which we use to force 'voice_note'.
    """
    # --- Voice note detection: filename always wins over mime type ---
    base_name = os.path.basename(filename).lower()
    if base_name.startswith("voice-"):
        return "voice_note"

    if not mime_type or mime_type == 'application/octet-stream':
        mime_type, _ = mimetypes.guess_type(filename)
        mime_type = mime_type or 'application/octet-stream'

    # --- Mime type checks ---
    if mime_type.startswith('image/'):
        return 'image'

    # audio/* covers mp3, wav, ogg, m4a, aac etc.
    if mime_type.startswith('audio/'):
        return 'audio'

    # video/webm that is NOT a voice note is a real video
    if mime_type.startswith('video/'):
        return 'video'

    if mime_type == 'application/pdf':
        return 'pdf'
    if mime_type in (
        'application/msword',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    ):
        return 'word'
    if mime_type in (
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ):
        return 'excel'

    # --- Extension fallback ---
    ext = os.path.splitext(filename)[1].lower()
    if ext in ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg'):
        return 'image'
    if ext in ('.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv'):
        return 'video'
    # .webm extension without a voice- prefix → treat as video
    if ext == '.webm':
        return 'video'
    if ext in ('.mp3', '.wav', '.ogg', '.m4a', '.aac'):
        return 'audio'
    if ext == '.pdf':
        return 'pdf'
    if ext in ('.doc', '.docx'):
        return 'word'
    if ext in ('.xls', '.xlsx'):
        return 'excel'

    return 'document'


def extract_metadata(media_file, file_obj):
    """
    Extract metadata based on file type.
    Uses only PIL for images (always available).
    Skips video/audio metadata gracefully if optional libs are missing
    or if the file is in-memory (no temporary_file_path).
    """
    try:
        ft = media_file.file_type

        # ── Images ──────────────────────────────────────────────────────
        if ft == 'image':
            file_obj.seek(0)
            img = Image.open(file_obj)
            media_file.dimensions = f"{img.width}x{img.height}"
            print(f"[extract_metadata] Image dimensions: {media_file.dimensions}")

            # Convert mode if needed (handles RGBA/P/LA PNGs)
            if img.mode in ('RGBA', 'P', 'LA'):
                img = img.convert('RGB')

            thumb = img.copy()
            thumb.thumbnail((200, 200))

            original_filename = os.path.basename(media_file.file.name)
            # Strip any extra dots from the base name for the thumbnail too
            name_parts   = original_filename.rsplit(".", 1)
            clean_base   = name_parts[0].replace(".", "_") if len(name_parts) > 1 else original_filename
            thumb_name   = f"{clean_base}.jpg" if len(name_parts) > 1 else f"{clean_base}_thumb.jpg"

            thumb_relative = f"thumbnails/{thumb_name}"
            thumb_absolute = os.path.join(settings.MEDIA_ROOT, thumb_relative)
            os.makedirs(os.path.dirname(thumb_absolute), exist_ok=True)
            thumb.save(thumb_absolute, format='JPEG', quality=85)

            media_file.thumbnail = thumb_relative
            print(f"[extract_metadata] Thumbnail saved: {thumb_relative}")

        # ── Video ────────────────────────────────────────────────────────
        elif ft == 'video':
            # Only attempt if moviepy is available AND the file is on disk
            try:
                from moviepy.editor import VideoFileClip
                # InMemoryUploadedFile has no temporary_file_path — skip gracefully
                if hasattr(file_obj, 'temporary_file_path'):
                    clip = VideoFileClip(file_obj.temporary_file_path())
                    media_file.duration   = clip.duration
                    media_file.dimensions = f"{clip.size[0]}x{clip.size[1]}"
                    clip.close()
                    print(f"[extract_metadata] Video duration: {media_file.duration}s | dims: {media_file.dimensions}")
                else:
                    print(f"[extract_metadata] Skipping video metadata — file is in-memory (no temp path)")
            except ImportError:
                print(f"[extract_metadata] moviepy not available — skipping video metadata")
            except Exception as ve:
                print(f"[extract_metadata] Video metadata error (non-fatal): {ve}")

        # ── Audio / Voice note ───────────────────────────────────────────
        elif ft in ('audio', 'voice_note'):
            try:
                from pydub import AudioSegment
                file_obj.seek(0)
                audio = AudioSegment.from_file(io.BytesIO(file_obj.read()))
                media_file.duration = len(audio) / 1000  # ms → seconds
                print(f"[extract_metadata] Audio duration: {media_file.duration}s")
            except ImportError:
                print(f"[extract_metadata] pydub not available — skipping audio metadata")
            except Exception as ae:
                print(f"[extract_metadata] Audio metadata error (non-fatal): {ae}")

        # ── PDF ──────────────────────────────────────────────────────────
        elif ft == 'pdf':
            try:
                import PyPDF2
                file_obj.seek(0)
                reader = PyPDF2.PdfReader(file_obj)
                media_file.page_count = len(reader.pages)
                print(f"[extract_metadata] PDF pages: {media_file.page_count}")
            except ImportError:
                print(f"[extract_metadata] PyPDF2 not available — skipping PDF metadata")
            except Exception as pe:
                print(f"[extract_metadata] PDF metadata error (non-fatal): {pe}")

        # ── Word ─────────────────────────────────────────────────────────
        elif ft == 'word':
            try:
                from docx import Document
                file_obj.seek(0)
                doc = Document(file_obj)
                media_file.word_count = sum(
                    len(p.text.split()) for p in doc.paragraphs
                )
                print(f"[extract_metadata] Word count: {media_file.word_count}")
            except ImportError:
                print(f"[extract_metadata] python-docx not available — skipping Word metadata")
            except Exception as we:
                print(f"[extract_metadata] Word metadata error (non-fatal): {we}")

        # ── Excel ────────────────────────────────────────────────────────
        elif ft == 'excel':
            try:
                from openpyxl import load_workbook
                file_obj.seek(0)
                wb = load_workbook(file_obj, read_only=True)
                media_file.page_count = len(wb.sheetnames)
                print(f"[extract_metadata] Excel sheets: {media_file.page_count}")
            except ImportError:
                print(f"[extract_metadata] openpyxl not available — skipping Excel metadata")
            except Exception as xe:
                print(f"[extract_metadata] Excel metadata error (non-fatal): {xe}")

        media_file.save()
        print(f"[extract_metadata] MediaFile ID={media_file.id} ({ft}) saved successfully")

    except Exception as e:
        logger.error(f"[extract_metadata] Unexpected error for {media_file.file_name}: {e}")
        print(f"[extract_metadata] ERROR: {e}")