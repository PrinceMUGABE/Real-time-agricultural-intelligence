# chatApp/views.py
# chatApp/views.py - Update the imports at the top
from datetime import timedelta
import logging
import os
from django.conf import settings
from django.utils.timezone import now
from django.db.models import Q, Count, Sum
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError, PermissionDenied
from django.http import FileResponse, Http404

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from userApp.models import CustomUser
from .models import (
    ChatRoom, ChatParticipant, Message, ChatRoomType,
    ChatRoomSettings, MessageVisibilityOverride, MediaFile, MediaFileType
)
from .serializers import (
    ChatRoomSerializer, MessageSerializer, ChatParticipantSerializer,
    CreateChatRoomSerializer, SendMessageSerializer, DeleteMessageSerializer,
    UpdateChatSettingsSerializer, BlockParticipantSerializer, UserBasicSerializer,
    AdminChatListSerializer, AdminParticipantDetailSerializer,
    AdminChatDetailSerializer, AdminUpdateParticipantRoleSerializer,
    AdminBlockParticipantSerializer, AdminChatStatsSerializer,
    MediaFileSerializer, EnhancedMessageSerializer, MessageFilterSerializer
)
from .translations import ct
from .media_helpers import determine_file_type, extract_metadata
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

import mimetypes
from django.db.models import Q, Count
from django.core.files.storage import default_storage
from django.http import FileResponse, Http404
from PIL import Image
from pydub import AudioSegment
import PyPDF2
from docx import Document
from openpyxl import load_workbook
import magic

from datetime import timedelta
import logging
import os
from django.utils.timezone import now
from django.db.models import Q, Count, Sum
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError, PermissionDenied
from django.http import FileResponse, Http404

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from userApp.models import CustomUser
from .models import (
    ChatRoom, ChatParticipant, Message, ChatRoomType,
    ChatRoomSettings, MessageVisibilityOverride
)

# Try to import MediaFile and MediaFileType - they might not exist yet
try:
    from .models import MediaFile, MediaFileType
    MEDIA_MODELS_AVAILABLE = True
except ImportError:
    MEDIA_MODELS_AVAILABLE = False
    MediaFile = None
    MediaFileType = None

from .serializers import (
    ChatRoomSerializer, MessageSerializer, ChatParticipantSerializer,
    CreateChatRoomSerializer, SendMessageSerializer, DeleteMessageSerializer,
    UpdateChatSettingsSerializer, BlockParticipantSerializer, UserBasicSerializer,
    AdminChatListSerializer, AdminParticipantDetailSerializer,
    AdminChatDetailSerializer, AdminUpdateParticipantRoleSerializer,
    AdminBlockParticipantSerializer, AdminChatStatsSerializer
)

# Try to import media serializers if they exist
try:
    from .serializers import (
        MediaFileSerializer, EnhancedMessageSerializer, MessageFilterSerializer
    )
    MEDIA_SERIALIZERS_AVAILABLE = True
except ImportError:
    MEDIA_SERIALIZERS_AVAILABLE = False
    MediaFileSerializer = None
    EnhancedMessageSerializer = None
    MessageFilterSerializer = None

from .translations import ct

# Try to import media_helpers
try:
    from .media_helpers import determine_file_type, extract_metadata, get_mime_type
    MEDIA_HELPERS_AVAILABLE = True
except ImportError:
    MEDIA_HELPERS_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("media_helpers not available - media upload disabled")

from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

logger = logging.getLogger(__name__)


# Conditional imports for optional dependencies
# Try to import moviepy
try:
    # Try different import paths for moviepy
    try:
        from moviepy.editor import VideoFileClip
        MOVIEPY_AVAILABLE = True
    except (ImportError, ModuleNotFoundError):
        try:
            from moviepy import VideoFileClip
            MOVIEPY_AVAILABLE = True
        except (ImportError, ModuleNotFoundError):
            MOVIEPY_AVAILABLE = False
            logger.warning("moviepy not available - video features disabled")
except Exception:
    MOVIEPY_AVAILABLE = False
    logger.warning("moviepy import failed - video features disabled")

# Try to import pydub
try:
    from pydub import AudioSegment
    PYDUB_AVAILABLE = True
except (ImportError, ModuleNotFoundError):
    PYDUB_AVAILABLE = False
    logger.warning("pydub not available - audio features disabled")

# Try to import PyPDF2
try:
    import PyPDF2
    PYPDF2_AVAILABLE = True
except (ImportError, ModuleNotFoundError):
    PYPDF2_AVAILABLE = False
    logger.warning("PyPDF2 not available - PDF features disabled")

# Try to import docx
try:
    from docx import Document
    DOCX_AVAILABLE = True
except (ImportError, ModuleNotFoundError):
    DOCX_AVAILABLE = False
    logger.warning("python-docx not available - Word features disabled")

# Try to import openpyxl
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except (ImportError, ModuleNotFoundError):
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel features disabled")


logger = logging.getLogger(__name__)


# Helper functions
def get_user_language(request):
    """Get user's preferred language"""
    if request.user.is_authenticated:
        return getattr(request.user, 'language', 'en')
    return request.headers.get('Accept-Language', 'en')[:2]


def handle_exception(e, context="", request=None):
    """Handle exceptions with translations"""
    lang = get_user_language(request) if request else 'en'
    logger.error(f"ERROR: {context}: {str(e)}")
    
    if isinstance(e, ValidationError):
        return Response(
            {'error': ct("validation_error", lang), 'details': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    elif isinstance(e, PermissionDenied):
        return Response(
            {'error': ct("permission_denied", lang), 'details': str(e)},
            status=status.HTTP_403_FORBIDDEN
        )
    else:
        return Response(
            {'error': ct("server_error", lang)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ====== Chat Room Management ======

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_my_chats(request):
    """Get all chat rooms the user belongs to"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        # Get user's chat rooms
        chat_rooms = ChatRoom.objects.filter(
            participants=user,
            is_active=True
        ).prefetch_related(
            'chat_participants__user',
            'messages'
        ).order_by('-updated_at')
        
        serializer = ChatRoomSerializer(
            chat_rooms, many=True, context={'request': request}
        )
        
        return Response({
            'success': True,
            'message': ct("chats_retrieved", lang),
            'count': chat_rooms.count(),
            'chats': serializer.data
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to get chats", request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_chat_room(request):
    """Create a new chat room"""
    print("Received request to create chat room with data:", request.data)
    lang = get_user_language(request)
    try:
        user = request.user
        chat_type = request.data.get('chat_type')
        user_id   = request.data.get('user_id')

        if not chat_type:
            return Response({
                'error': ct("chat_type_required", lang)
            }, status=status.HTTP_400_BAD_REQUEST)

        if chat_type not in [ct_val for ct_val, _ in ChatRoomType.choices]:
            return Response({
                'error': ct("invalid_chat_type", lang)
            }, status=status.HTTP_400_BAD_REQUEST)

        # ── One-on-one chat ──────────────────────────────────────────────
        if chat_type == ChatRoomType.ONE_ON_ONE:
            if not user_id:
                return Response({
                    'error': "user_id is required for one-on-one chats"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Resolve target user — only check is_active, no status string filter
            try:
                other_user = CustomUser.objects.get(id=user_id, is_active=True)
            except CustomUser.DoesNotExist:
                return Response({
                    'error': "User not found or inactive"
                }, status=status.HTTP_404_NOT_FOUND)

            if other_user.id == user.id:
                return Response({
                    'error': "You cannot start a chat with yourself"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Return existing chat if one already exists between these two users
            existing = ChatRoom.objects.filter(
                chat_type=ChatRoomType.ONE_ON_ONE,
                user1__in=[user, other_user],
                user2__in=[user, other_user]
            ).first()

            if existing:
                print(f"Returning existing one-on-one chat ID={existing.id}")
                return Response({
                    'success': True,
                    'message': ct("chat_exists", lang),
                    'chat': ChatRoomSerializer(existing, context={'request': request}).data
                })

            # Create new one-on-one chat
            chat_room = ChatRoom.objects.create(
                chat_type=chat_type,
                user1=user,
                user2=other_user,
                created_by=user
            )
            print(f"Created one-on-one chat ID={chat_room.id}")

            # Add both users as participants
            chat_role  = 'admin' if user.role == 'admin' else 'member'
            other_role = 'admin' if other_user.role == 'admin' else 'member'
            ChatParticipant.objects.create(chat_room=chat_room, user=user,       role=chat_role)
            ChatParticipant.objects.create(chat_room=chat_room, user=other_user, role=other_role)

            # Add all admins as silent observers
            admins = CustomUser.objects.filter(role='admin', is_active=True).exclude(
                id__in=[user.id, other_user.id]
            )
            for admin in admins:
                ChatParticipant.objects.create(chat_room=chat_room, user=admin, role='observer')

        # ── System chats (global / farmers / buyers) ─────────────────────
        else:
            if user.role != 'admin':
                return Response({
                    'error': ct("admin_required", lang)
                }, status=status.HTTP_403_FORBIDDEN)

            existing = ChatRoom.objects.filter(chat_type=chat_type).first()
            if existing:
                print(f"Returning existing system chat type={chat_type} ID={existing.id}")
                return Response({
                    'success': True,
                    'message': ct("chat_exists", lang),
                    'chat': ChatRoomSerializer(existing, context={'request': request}).data
                })

            chat_room = ChatRoom.objects.create(chat_type=chat_type, created_by=user)
            print(f"Created system chat type={chat_type} ID={chat_room.id}")

            # Add appropriate users based on chat type
            if chat_type == ChatRoomType.GLOBAL:
                participants = CustomUser.objects.filter(is_active=True)
            elif chat_type == ChatRoomType.FARMERS:
                participants = CustomUser.objects.filter(
                    Q(role='admin') | Q(role='farmer'),
                    is_active=True
                )
            else:  # buyers
                participants = CustomUser.objects.filter(
                    Q(role='admin') | Q(role='buyer'),
                    is_active=True
                )

            for participant in participants:
                role = 'admin' if participant.role == 'admin' else 'member'
                ChatParticipant.objects.create(chat_room=chat_room, user=participant, role=role)

        # Create default settings
        ChatRoomSettings.objects.get_or_create(chat_room=chat_room)

        return Response({
            'success': True,
            'message': ct("chat_created", lang),
            'chat': ChatRoomSerializer(chat_room, context={'request': request}).data
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        print(f"Exception in create_chat_room: {e}")
        return handle_exception(e, "Failed to create chat", request)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_chat_room(request, room_id):
    """Get specific chat room with messages"""
    lang = get_user_language(request)
    try:
        user = request.user
        chat_room = get_object_or_404(ChatRoom, id=room_id, is_active=True)
        
        # Check access
        is_participant = chat_room.participants.filter(id=user.id).exists()
        if not is_participant and user.role != 'admin':
            return Response({
                'error': ct("not_participant", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check if user is blocked
        participant = chat_room.chat_participants.filter(user=user).first()
        if participant and participant.is_blocked:
            return Response({
                'error': ct("user_blocked", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get messages with pagination
        limit = int(request.GET.get('limit', 50))
        offset = int(request.GET.get('offset', 0))
        
        messages = chat_room.messages.filter(is_deleted=False).order_by('created_at')
        
        # Filter messages by visibility
        visible_messages = []
        for msg in messages[offset:offset + limit]:
            if msg.can_view(user):
                visible_messages.append(msg)
        
        message_serializer = MessageSerializer(
            visible_messages, many=True, context={'request': request}
        )
        
        # Update last read time
        if is_participant:
            ChatParticipant.objects.filter(
                chat_room=chat_room,
                user=user
            ).update(last_read_at=now())
        
        return Response({
            'success': True,
            'chat': ChatRoomSerializer(
                chat_room, context={'request': request}
            ).data,
            'messages': message_serializer.data,
            'pagination': {
                'total': messages.count(),
                'limit': limit,
                'offset': offset,
                'has_more': (offset + limit) < messages.count()
            }
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to get chat room", request)


# ====== Message Handling ======

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_message(request):
    """Send a message to a chat room"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        serializer = SendMessageSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'error': ct("invalid_message", lang),
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        chat_room = get_object_or_404(
            ChatRoom, id=data['chat_room_id'], is_active=True
        )
        
        # Check if user can send messages
        if not chat_room.can_user_send_message(user):
            return Response({
                'error': ct("cannot_send_message", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Create message
        message = Message.objects.create(
            chat_room=chat_room,
            sender=user,
            message_type=data['message_type'],
            content=data.get('content', ''),
            attachment=data.get('attachment')
        )
        
        # Update chat room timestamp
        chat_room.updated_at = now()
        chat_room.save()
        
        # Broadcast via WebSocket
        try:
            channel_layer = get_channel_layer()
            message_data = MessageSerializer(
                message, context={'request': request}
            ).data
            
            # Send to all participants (visibility handled client-side)
            async_to_sync(channel_layer.group_send)(
                f"chat_{chat_room.id}",
                {
                    'type': 'chat_message',
                    'message': message_data
                }
            )
        except Exception as e:
            logger.error(f"WebSocket broadcast error: {e}")
        
        return Response({
            'success': True,
            'message': ct("message_sent", lang),
            'data': MessageSerializer(
                message, context={'request': request}
            ).data
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return handle_exception(e, "Failed to send message", request)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_message(request, message_id):
    """Delete a message"""
    lang = get_user_language(request)
    try:
        user = request.user
        message = get_object_or_404(Message, id=message_id)
        
        # Check if user can delete
        if not message.can_delete(user):
            return Response({
                'error': ct("cannot_delete", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get delete type
        delete_type = request.data.get('delete_type', 'for_me')
        
        # Perform deletion
        message.delete_for_user(user, delete_type)
        
        # Broadcast deletion
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"chat_{message.chat_room.id}",
                {
                    'type': 'message_deleted',
                    'message_id': message_id,
                    'deleted_by': user.id,
                    'delete_type': delete_type
                }
            )
        except Exception as e:
            logger.error(f"WebSocket broadcast error: {e}")
        
        return Response({
            'success': True,
            'message': ct("message_deleted", lang)
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to delete message", request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_message_read(request, message_id):
    """Mark a message as read"""
    lang = get_user_language(request)
    try:
        user = request.user
        message = get_object_or_404(Message, id=message_id)
        
        # Check if user can view message
        if not message.can_view(user):
            return Response({
                'error': ct("cannot_view_message", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Mark as read
        message.read_by.add(user)
        
        return Response({
            'success': True,
            'message': ct("message_marked_read", lang)
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to mark message read", request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_all_read(request, room_id):
    """Mark all messages in a chat as read"""
    lang = get_user_language(request)
    try:
        user = request.user
        chat_room = get_object_or_404(ChatRoom, id=room_id, is_active=True)
        
        # Update last read time
        ChatParticipant.objects.filter(
            chat_room=chat_room,
            user=user
        ).update(last_read_at=now())
        
        return Response({
            'success': True,
            'message': ct("all_marked_read", lang)
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to mark all read", request)


# ====== Admin Management ======

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_chat_settings(request, room_id):
    """Update chat room settings (admin only)"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id, is_active=True)
        
        serializer = UpdateChatSettingsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'error': ct("invalid_settings", lang),
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get or create settings
        settings, _ = ChatRoomSettings.objects.get_or_create(chat_room=chat_room)
        settings.allowed_senders = serializer.validated_data['allowed_senders']
        settings.updated_by = user
        settings.save()
        
        return Response({
            'success': True,
            'message': ct("settings_updated", lang),
            'settings': {
                'allowed_senders': settings.allowed_senders
            }
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to update settings", request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def manage_participant(request, room_id):
    """Add or remove participant (admin only)"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id, is_active=True)
        
        action = request.data.get('action')
        target_user_id = request.data.get('user_id')
        
        if not action or not target_user_id:
            return Response({
                'error': ct("action_user_required", lang)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        target_user = get_object_or_404(CustomUser, id=target_user_id)
        
        if action == 'add':
            # Add participant
            role = request.data.get('role', 'member')
            ChatParticipant.objects.get_or_create(
                chat_room=chat_room,
                user=target_user,
                defaults={'role': role}
            )
            message = ct("participant_added", lang)
            
        elif action == 'remove':
            # Remove participant
            ChatParticipant.objects.filter(
                chat_room=chat_room,
                user=target_user
            ).delete()
            message = ct("participant_removed", lang)
            
        else:
            return Response({
                'error': ct("invalid_action", lang)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to manage participant", request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def block_participant(request, room_id):
    """Block or unblock a participant (admin only)"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id, is_active=True)
        
        serializer = BlockParticipantSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'error': ct("invalid_data", lang),
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        target_user = get_object_or_404(CustomUser, id=data['user_id'])
        
        participant = get_object_or_404(
            ChatParticipant,
            chat_room=chat_room,
            user=target_user
        )
        
        if data['action'] == 'block':
            participant.is_blocked = True
            participant.blocked_at = now()
            participant.blocked_by = user
            message = ct("user_blocked", lang)
        else:
            participant.is_blocked = False
            participant.blocked_at = None
            participant.blocked_by = None
            message = ct("user_unblocked", lang)
        
        participant.save()
        
        return Response({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to block/unblock participant", request)


# ====== Statistics ======

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_unread_counts(request):
    """Get unread message counts for all user's chats"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        chats = ChatRoom.objects.filter(
            participants=user,
            is_active=True
        )
        
        result = []
        for chat in chats:
            try:
                participant = ChatParticipant.objects.get(
                    chat_room=chat, user=user
                )
                
                # Count unread messages visible to user
                messages = chat.messages.filter(
                    is_deleted=False
                ).exclude(sender=user)
                
                if participant.last_read_at:
                    messages = messages.filter(
                        created_at__gt=participant.last_read_at
                    )
                
                # Filter by visibility
                unread_count = 0
                for msg in messages:
                    if msg.can_view(user):
                        unread_count += 1
                
                if unread_count > 0:
                    result.append({
                        'chat_id': chat.id,
                        'chat_name': chat.name,
                        'chat_type': chat.chat_type,
                        'unread_count': unread_count
                    })
                    
            except ChatParticipant.DoesNotExist:
                continue
        
        return Response({
            'success': True,
            'total_unread': sum(item['unread_count'] for item in result),
            'chats': result
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to get unread counts", request)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_chat_stats(request, room_id):
    """Get statistics for a specific chat room"""
    lang = get_user_language(request)
    try:
        user = request.user
        chat_room = get_object_or_404(ChatRoom, id=room_id, is_active=True)
        
        # Check access
        if not chat_room.participants.filter(id=user.id).exists() and user.role != 'admin':
            return Response({
                'error': ct("not_participant", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Basic stats
        total_messages = chat_room.messages.filter(is_deleted=False).count()
        
        # Messages per user
        messages_by_user = chat_room.messages.filter(
            is_deleted=False
        ).values('sender__full_name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Active participants
        active_participants = chat_room.chat_participants.filter(
            is_blocked=False
        ).count()
        
        stats = {
            'total_messages': total_messages,
            'active_participants': active_participants,
            'created_at': chat_room.created_at,
            'last_message_at': chat_room.updated_at,
            'messages_by_user': list(messages_by_user)
        }
        
        # Admin-only stats
        if user.role == 'admin':
            stats['blocked_users'] = chat_room.chat_participants.filter(
                is_blocked=True
            ).count()
            stats['total_participants'] = chat_room.participants.count()
        
        return Response({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to get chat stats", request)
    
    


# ====== ADMIN CHAT MANAGEMENT ======

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_get_all_chats(request):
    """Admin: Get all chats with statistics and pagination"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        # Check if user is admin
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get pagination parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        sort_by = request.GET.get('sort_by', 'updated_at')
        sort_dir = request.GET.get('sort_dir', 'desc')
        
        # Get filter parameters
        chat_type = request.GET.get('chat_type')
        status_filter = request.GET.get('status')
        search = request.GET.get('search')
        participant_id = request.GET.get('participant_id')  # New filter
        
        # Build query
        queryset = ChatRoom.objects.all()
        
        # Apply filters
        if chat_type:
            queryset = queryset.filter(chat_type=chat_type)
        
        if status_filter:
            is_active = status_filter == 'active'
            queryset = queryset.filter(is_active=is_active)
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(user1__full_name__icontains=search) |
                Q(user2__full_name__icontains=search)
            )
        
        # Filter by participant if specified
        if participant_id:
            queryset = queryset.filter(
                Q(user1_id=participant_id) | 
                Q(user2_id=participant_id) |
                Q(participants__id=participant_id)
            ).distinct()
        
        # Apply sorting
        sort_prefix = '' if sort_dir == 'asc' else '-'
        if sort_by == 'name':
            queryset = queryset.order_by(f'{sort_prefix}name')
        elif sort_by == 'chat_type':
            queryset = queryset.order_by(f'{sort_prefix}chat_type')
        elif sort_by == 'created_at':
            queryset = queryset.order_by(f'{sort_prefix}created_at')
        elif sort_by == 'updated_at':
            queryset = queryset.order_by(f'{sort_prefix}updated_at')
        elif sort_by == 'participant_count':
            # Annotate with participant count and sort
            queryset = queryset.annotate(
                participant_count=Count('participants')
            ).order_by(f'{sort_prefix}participant_count')
        
        # Calculate pagination
        total = queryset.count()
        total_pages = (total + page_size - 1) // page_size
        start = (page - 1) * page_size
        end = start + page_size
        
        # Get paginated results
        chats = queryset[start:end]
        
        # Calculate statistics
        stats = {
            'total_chats': ChatRoom.objects.count(),
            'active_chats': ChatRoom.objects.filter(is_active=True).count(),
            'inactive_chats': ChatRoom.objects.filter(is_active=False).count(),
            'global_chats': ChatRoom.objects.filter(chat_type=ChatRoomType.GLOBAL).count(),
            'farmers_chats': ChatRoom.objects.filter(chat_type=ChatRoomType.FARMERS).count(),
            'buyers_chats': ChatRoom.objects.filter(chat_type=ChatRoomType.BUYERS).count(),
            'one_on_one_chats': ChatRoom.objects.filter(chat_type=ChatRoomType.ONE_ON_ONE).count(),
            'total_participants': ChatParticipant.objects.count(),
            'total_messages': Message.objects.filter(is_deleted=False).count(),
        }
        
        # Serialize chats with message counts
        chat_data = []
        for chat in chats:
            chat_dict = ChatRoomSerializer(chat, context={'request': request}).data
            chat_dict['message_count'] = chat.messages.filter(is_deleted=False).count()
            chat_dict['participants'] = ChatParticipantSerializer(
                chat.chat_participants.select_related('user'), 
                many=True, 
                context={'request': request}
            ).data
            chat_data.append(chat_dict)
        
        return Response({
            'success': True,
            'chats': chat_data,
            'total': total,
            'total_pages': total_pages,
            'page': page,
            'page_size': page_size,
            'stats': stats
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to fetch chats", request)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_create_chat(request):
    print(f"Received admin request to create chat with data: {request.data}")
    """Admin: Create a new chat room (system chats or one-on-one)"""
    lang = get_user_language(request)
    try:
        user = request.user
        print(f"Authenticated user: {user.full_name} (Role: {user.role})")

        # Check if user is admin
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)

        chat_type = request.data.get('chat_type')
        name = request.data.get('name', '')
        user_id = request.data.get('user_id')  # For one-on-one chats

        if not chat_type:
            return Response({
                'error': ct("chat_type_required", lang)
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validate chat type
        valid_types = [ChatRoomType.GLOBAL, ChatRoomType.FARMERS,
                       ChatRoomType.BUYERS, ChatRoomType.ONE_ON_ONE]

        if chat_type not in valid_types:
            return Response({
                'error': ct("invalid_chat_type", lang)
            }, status=status.HTTP_400_BAD_REQUEST)

        # --- Check if system chat already exists, return it immediately ---
        if chat_type in [ChatRoomType.GLOBAL, ChatRoomType.FARMERS, ChatRoomType.BUYERS]:
            existing = ChatRoom.objects.filter(chat_type=chat_type).first()
            if existing:
                # Ensure settings exist for this room (backfill if missing)
                ChatRoomSettings.objects.get_or_create(chat_room=existing)
                return Response({
                    'success': True,
                    'message': ct("chat_exists", lang),
                    'chat': ChatRoomSerializer(existing, context={'request': request}).data
                })

        # --- Resolve one-on-one user BEFORE creating the room ---
        other_user = None
        if chat_type == ChatRoomType.ONE_ON_ONE:
            if not user_id:
                return Response({
                    'error': ct("user_id_required", lang)
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                other_user = CustomUser.objects.get(id=user_id, is_active=True)
            except CustomUser.DoesNotExist:
                return Response({
                    'error': ct("user_not_found", lang)
                }, status=status.HTTP_404_NOT_FOUND)

            # Prevent duplicate one-on-one chats between the same two users
            existing_one_on_one = ChatRoom.objects.filter(
                chat_type=ChatRoomType.ONE_ON_ONE,
                user1=user,
                user2=other_user
            ).first() or ChatRoom.objects.filter(
                chat_type=ChatRoomType.ONE_ON_ONE,
                user1=other_user,
                user2=user
            ).first()

            if existing_one_on_one:
                ChatRoomSettings.objects.get_or_create(chat_room=existing_one_on_one)
                return Response({
                    'success': True,
                    'message': ct("chat_exists", lang),
                    'chat': ChatRoomSerializer(existing_one_on_one, context={'request': request}).data
                })

        # --- Create chat room with user1/user2 included from the start ---
        chat_room = ChatRoom.objects.create(
            chat_type=chat_type,
            name=name,
            created_by=user,
            is_active=True,
            user1=user if chat_type == ChatRoomType.ONE_ON_ONE else None,
            user2=other_user if chat_type == ChatRoomType.ONE_ON_ONE else None,
        )

        # --- Add participants based on chat type ---
        if chat_type == ChatRoomType.ONE_ON_ONE:
            ChatParticipant.objects.create(chat_room=chat_room, user=user, role='admin')
            ChatParticipant.objects.create(chat_room=chat_room, user=other_user, role='member')

            # Add all other admins as observers
            admins = CustomUser.objects.filter(role='admin', is_active=True).exclude(
                id__in=[user.id, other_user.id]
            )
            for admin in admins:
                ChatParticipant.objects.create(chat_room=chat_room, user=admin, role='observer')

        elif chat_type == ChatRoomType.GLOBAL:
            users = CustomUser.objects.filter(is_active=True)
            for participant in users:
                role = 'admin' if participant.role == 'admin' else 'member'
                ChatParticipant.objects.create(chat_room=chat_room, user=participant, role=role)

        elif chat_type == ChatRoomType.FARMERS:
            users = CustomUser.objects.filter(
                Q(role='admin') | Q(role='farmer'),
                is_active=True
            )
            for participant in users:
                role = 'admin' if participant.role == 'admin' else 'member'
                ChatParticipant.objects.create(chat_room=chat_room, user=participant, role=role)

        elif chat_type == ChatRoomType.BUYERS:
            users = CustomUser.objects.filter(
                Q(role='admin') | Q(role='buyer'),
                is_active=True
            )
            for participant in users:
                role = 'admin' if participant.role == 'admin' else 'member'
                ChatParticipant.objects.create(chat_room=chat_room, user=participant, role=role)

        # --- Create default settings (get_or_create prevents duplicate key errors) ---
        ChatRoomSettings.objects.get_or_create(chat_room=chat_room)

        return Response({
            'success': True,
            'message': ct("chat_created", lang),
            'chat': ChatRoomSerializer(chat_room, context={'request': request}).data
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return handle_exception(e, "Failed to create chat", request)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_get_chat_details(request, room_id):
    """Admin: Get detailed information about a specific chat"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        # Check if user is admin
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id)
        
        # Get participants with details
        participants = chat_room.chat_participants.select_related('user', 'blocked_by').all()
        participant_data = []
        
        for p in participants:
            participant_data.append({
                'id': p.user.id,
                'full_name': p.user.full_name,
                'email': p.user.email,
                'phone': p.user.phone_number,
                'role': p.role,
                'is_blocked': p.is_blocked,
                'blocked_at': p.blocked_at,
                'blocked_by': p.blocked_by.full_name if p.blocked_by else None,
                'is_muted': p.is_muted,
                'joined_at': p.joined_at,
                'last_read_at': p.last_read_at
            })
        
        # Get message statistics
        total_messages = chat_room.messages.filter(is_deleted=False).count()
        messages_last_7_days = chat_room.messages.filter(
            is_deleted=False,
            created_at__gte=now() - timedelta(days=7)
        ).count()
        
        # Get top contributors
        top_contributors = chat_room.messages.filter(
            is_deleted=False
        ).values('sender__full_name').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        
        # Get settings
        try:
            settings = chat_room.settings
            settings_data = {
                'allowed_senders': settings.allowed_senders,
                'updated_at': settings.updated_at,
                'updated_by': settings.updated_by.full_name if settings.updated_by else None
            }
        except ChatRoomSettings.DoesNotExist:
            settings_data = None
        
        return Response({
            'success': True,
            'chat': {
                'id': chat_room.id,
                'name': chat_room.name,
                'chat_type': chat_room.chat_type,
                'is_active': chat_room.is_active,
                'created_at': chat_room.created_at,
                'updated_at': chat_room.updated_at,
                'created_by': chat_room.created_by.full_name if chat_room.created_by else None,
                'user1': UserBasicSerializer(chat_room.user1).data if chat_room.user1 else None,
                'user2': UserBasicSerializer(chat_room.user2).data if chat_room.user2 else None
            },
            'participants': participant_data,
            'statistics': {
                'total_participants': len(participant_data),
                'blocked_participants': sum(1 for p in participant_data if p['is_blocked']),
                'total_messages': total_messages,
                'messages_last_7_days': messages_last_7_days,
                'top_contributors': list(top_contributors)
            },
            'settings': settings_data
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to get chat details", request)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def admin_update_chat_settings(request, room_id):
    """Admin: Update chat room settings"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        # Check if user is admin
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id)
        
        name = request.data.get('name')
        allowed_senders = request.data.get('allowed_senders')
        is_active = request.data.get('is_active')
        
        # Update chat room
        if name is not None:
            chat_room.name = name
        if is_active is not None:
            chat_room.is_active = is_active
        chat_room.save()
        
        # Update settings
        if allowed_senders:
            settings, created = ChatRoomSettings.objects.get_or_create(chat_room=chat_room)
            settings.allowed_senders = allowed_senders
            settings.updated_by = user
            settings.save()
        
        return Response({
            'success': True,
            'message': ct("settings_updated", lang),
            'chat': ChatRoomSerializer(chat_room, context={'request': request}).data
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to update settings", request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_update_participant_role(request, room_id):
    """Admin: Update a participant's role in a chat"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        # Check if user is admin
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id)
        user_id = request.data.get('user_id')
        role = request.data.get('role')
        
        if not user_id or not role:
            return Response({
                'error': ct("user_id_role_required", lang)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate role
        valid_roles = ['admin', 'member', 'observer']
        if role not in valid_roles:
            return Response({
                'error': ct("invalid_role", lang)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        participant = get_object_or_404(
            ChatParticipant,
            chat_room=chat_room,
            user_id=user_id
        )
        
        participant.role = role
        participant.save()
        
        # Broadcast role change via WebSocket
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"chat_{chat_room.id}",
                {
                    'type': 'participant_role_updated',
                    'user_id': user_id,
                    'new_role': role,
                    'updated_by': user.id
                }
            )
        except Exception as e:
            logger.error(f"WebSocket broadcast error: {e}")
        
        return Response({
            'success': True,
            'message': ct("participant_role_updated", lang)
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to update participant role", request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_block_participant(request, room_id):
    """Admin: Block or unblock a participant in a chat"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        # Check if user is admin
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id)
        user_id = request.data.get('user_id')
        block = request.data.get('block', True)
        
        if not user_id:
            return Response({
                'error': ct("user_id_required", lang)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        participant = get_object_or_404(
            ChatParticipant,
            chat_room=chat_room,
            user_id=user_id
        )
        
        if block:
            participant.is_blocked = True
            participant.blocked_at = now()
            participant.blocked_by = user
            message_key = "user_blocked"
        else:
            participant.is_blocked = False
            participant.blocked_at = None
            participant.blocked_by = None
            message_key = "user_unblocked"
        
        participant.save()
        
        # Broadcast block status via WebSocket
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"chat_{chat_room.id}",
                {
                    'type': 'participant_blocked',
                    'user_id': user_id,
                    'is_blocked': block,
                    'updated_by': user.id
                }
            )
        except Exception as e:
            logger.error(f"WebSocket broadcast error: {e}")
        
        return Response({
            'success': True,
            'message': ct(message_key, lang)
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to block/unblock participant", request)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def admin_delete_chat(request, room_id):
    """Admin: Permanently delete a chat room and all its messages"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        # Check if user is admin
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id)
        
        # Store chat info for response
        chat_info = {
            'id': chat_room.id,
            'name': chat_room.name,
            'chat_type': chat_room.chat_type
        }
        
        # Delete the chat room (cascades to messages, participants, settings)
        chat_room.delete()
        
        return Response({
            'success': True,
            'message': ct("chat_deleted", lang),
            'chat': chat_info
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to delete chat", request)
    
    
    

# ====== MEDIA FILE MANAGEMENT ======

import re
from django.utils.text import slugify

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_media(request):
    """Upload media files to a chat room"""
    lang = get_user_language(request)
    try:
        user = request.user
        chat_room_id = request.data.get('chat_room_id')
        files = request.FILES.getlist('files')

        print(f"[upload_media] User: {user.full_name} (Role: {user.role}) | Room ID: {chat_room_id}")
        print(f"[upload_media] Files received: {[f.name for f in files]}")

        if not chat_room_id:
            return Response({'error': ct("chat_room_id_required", lang)}, status=status.HTTP_400_BAD_REQUEST)

        if not files:
            return Response({'error': ct("no_files_uploaded", lang)}, status=status.HTTP_400_BAD_REQUEST)

        chat_room = get_object_or_404(ChatRoom, id=chat_room_id, is_active=True)
        print(f"[upload_media] Chat room: '{chat_room.name}' (Type: {chat_room.chat_type})")

        if not chat_room.can_user_send_message(user):
            print(f"[upload_media] Permission denied for {user.full_name}")
            return Response({'error': ct("cannot_send_message", lang)}, status=status.HTTP_403_FORBIDDEN)

        # Create message first so FK is available for each MediaFile
        message = Message.objects.create(
            chat_room=chat_room,
            sender=user,
            message_type='file',
            content=request.data.get('content', ''),
            has_media=True
        )
        print(f"[upload_media] Message created: ID={message.id}")

        uploaded_media = []

        for file in files:
            print(f"[upload_media] Processing: '{file.name}' | Size: {file.size} bytes")

            # --- Sanitize filename: remove problematic extensions mid-name ---
            # e.g. "UseCase_Diagram.drawio.png" → "UseCase_Diagram_drawio.png"
            original_name = file.name
            name_parts    = original_name.rsplit(".", 1)           # split off final extension only
            base          = name_parts[0] if len(name_parts) > 1 else original_name
            ext           = f".{name_parts[1]}" if len(name_parts) > 1 else ""
            # Replace any dots in the base with underscores so storage doesn't get confused
            clean_base    = base.replace(".", "_")
            # Also strip non-ASCII / special chars that confuse file systems
            clean_base    = re.sub(r'[^\w\-]', '_', clean_base)
            file.name     = f"{clean_base}{ext}"
            print(f"[upload_media] Sanitized filename: '{original_name}' → '{file.name}'")

            # Step 1: detect mime, then reset pointer
            file.seek(0)
            mime_type = get_mime_type(file)
            file_type = determine_file_type(file.name, mime_type)
            print(f"[upload_media] mime_type: {mime_type} | file_type: {file_type}")

            # Step 2: reset before Django writes to disk
            file.seek(0)
            media_file = MediaFile.objects.create(
                message=message,
                file=file,
                file_name=original_name,   # keep human-readable original name for display
                file_size=file.size,
                mime_type=mime_type,
                file_type=file_type,
                uploaded_by=user
            )
            print(f"[upload_media] MediaFile ID={media_file.id} | saved path: {media_file.file.name}")

            # Step 3: reset before metadata extraction
            file.seek(0)
            extract_metadata(media_file, file)

            # Keep M2M in sync for backward compatibility
            message.media_files.add(media_file)
            uploaded_media.append(media_file)
            print(f"[upload_media] ✓ File saved successfully: {media_file.file.name}")

        print(f"[upload_media] Total uploaded: {len(uploaded_media)} file(s)")

        chat_room.updated_at = now()
        chat_room.save()

        try:
            channel_layer = get_channel_layer()
            message_data  = EnhancedMessageSerializer(message, context={'request': request}).data
            async_to_sync(channel_layer.group_send)(
                f"chat_{chat_room.id}",
                {'type': 'chat_message', 'message': message_data}
            )
            print(f"[upload_media] WebSocket broadcast sent for message ID={message.id}")
        except Exception as ws_err:
            logger.error(f"[upload_media] WebSocket error: {ws_err}")

        return Response({
            'success': True,
            'message': ct("media_uploaded", lang),
            'data': EnhancedMessageSerializer(message, context={'request': request}).data
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return handle_exception(e, "Failed to upload media", request)    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_chat_media(request, room_id):
    """Get all media files in a chat room with filtering"""
    lang = get_user_language(request)

    if not MEDIA_MODELS_AVAILABLE:
        return Response({
            'error': 'Media features are not available',
            'details': 'Please run migrations first'
        }, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    try:
        user = request.user
        print(f"[get_chat_media] User: {user.full_name} (Role: {user.role}) | Room ID: {room_id}")

        chat_room = get_object_or_404(ChatRoom, id=room_id, is_active=True)
        print(f"[get_chat_media] Chat room found: '{chat_room.name}' (Type: {chat_room.chat_type})")

        # Check access
        is_participant = chat_room.participants.filter(id=user.id).exists()
        if not is_participant and user.role != 'admin':
            print(f"[get_chat_media] Access denied — user {user.full_name} is not a participant")
            return Response({
                'error': ct("not_participant", lang)
            }, status=status.HTTP_403_FORBIDDEN)

        # Get filter parameters
        media_type = request.GET.get('media_type')
        file_extension = request.GET.get('file_extension')
        sender_id = request.GET.get('sender_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        print(f"[get_chat_media] Filters — media_type: {media_type}, file_extension: {file_extension}, "
              f"sender_id: {sender_id}, start_date: {start_date}, end_date: {end_date}")

        # Build base message query
        messages = chat_room.messages.filter(is_deleted=False, has_media=True)
        print(f"[get_chat_media] Messages with media (before visibility filter): {messages.count()}")

        # Filter messages by visibility for non-admin users
        if user.role != 'admin':
            visible_message_ids = [msg.id for msg in messages if msg.can_view(user)]
            messages = messages.filter(id__in=visible_message_ids)
            print(f"[get_chat_media] Visible messages after permission filter: {messages.count()}")

        # --- Query via M2M 'messages' reverse relation (covers both old and new uploads) ---
        media_files = MediaFile.objects.filter(
            messages__in=messages       # 'messages' is the M2M related_name on MediaFile
        ).select_related('uploaded_by').distinct()

        print(f"[get_chat_media] Total media files before filters: {media_files.count()}")

        # Diagnose FK vs M2M state for transparency
        fk_count = MediaFile.objects.filter(message__in=messages).count()
        m2m_count = MediaFile.objects.filter(messages__in=messages).distinct().count()
        print(f"[get_chat_media] Diagnostic — via FK: {fk_count} | via M2M: {m2m_count}")

        # Apply filters
        if media_type:
            media_files = media_files.filter(file_type=media_type)
            print(f"[get_chat_media] After media_type filter '{media_type}': {media_files.count()}")

        if file_extension:
            media_files = media_files.filter(file_name__icontains=f'.{file_extension}')
            print(f"[get_chat_media] After file_extension filter '.{file_extension}': {media_files.count()}")

        if sender_id:
            media_files = media_files.filter(uploaded_by_id=sender_id)
            print(f"[get_chat_media] After sender_id filter '{sender_id}': {media_files.count()}")

        if start_date:
            media_files = media_files.filter(uploaded_at__gte=start_date)
            print(f"[get_chat_media] After start_date filter '{start_date}': {media_files.count()}")

        if end_date:
            media_files = media_files.filter(uploaded_at__lte=end_date)
            print(f"[get_chat_media] After end_date filter '{end_date}': {media_files.count()}")

        # Calculate statistics
        total = media_files.count()
        total_size = media_files.aggregate(total=Sum('file_size'))['total'] or 0

        type_stats = media_files.values('file_type').annotate(
            count=Count('id'),
            total_size=Sum('file_size')
        ).order_by('file_type')

        formatted_type_stats = []
        for stat in type_stats:
            stat_dict = dict(stat)
            size = stat_dict['total_size'] or 0
            for unit in ['B', 'KB', 'MB', 'GB']:
                if size < 1024.0:
                    stat_dict['total_size_human'] = f"{size:.1f} {unit}"
                    break
                size /= 1024.0
            formatted_type_stats.append(stat_dict)

        # Format total size
        total_size_human = "0 B"
        size = total_size
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024.0:
                total_size_human = f"{size:.1f} {unit}"
                break
            size /= 1024.0

        print(f"[get_chat_media] Final result — {total} file(s), total size: {total_size_human}")
        print(f"[get_chat_media] Breakdown by type: { {s['file_type']: s['count'] for s in formatted_type_stats} }")

        # Serialize and return
        media_files_ordered = media_files.order_by('-uploaded_at')
        serializer = MediaFileSerializer(
            media_files_ordered, many=True, context={'request': request}
        )

        return Response({
            'success': True,
            'media_files': serializer.data,
            'statistics': {
                'total': total,
                'total_size': total_size,
                'total_size_human': total_size_human,
                'by_type': formatted_type_stats
            }
        })

    except Exception as e:
        return handle_exception(e, "Failed to get media files", request)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_chat_messages(request, room_id):
    """Get all messages in a chat room"""
    lang = get_user_language(request)
    try:
        user = request.user
        chat_room = get_object_or_404(ChatRoom, id=room_id, is_active=True)

        # Check access
        is_participant = chat_room.participants.filter(id=user.id).exists()
        if not is_participant and user.role != 'admin':
            return Response({
                'error': ct("not_participant", lang)
            }, status=status.HTTP_403_FORBIDDEN)

        # Check if user is blocked
        participant = chat_room.chat_participants.filter(user=user).first()
        if participant and participant.is_blocked:
            return Response({
                'error': ct("user_blocked_error", lang)
            }, status=status.HTTP_403_FORBIDDEN)

        print(f"[get_chat_messages] User: {user.full_name} | Room: {chat_room.name} (ID={room_id})")

        # Fetch all non-deleted messages with both media relationships prefetched
        messages_qs = chat_room.messages.filter(
            is_deleted=False
        ).prefetch_related(
            'media_files',              # M2M: message.media_files.all()
            'media_file',               # FK reverse: message.media_file.all()
            'media_files__uploaded_by',
            'media_file__uploaded_by',
        ).select_related(
            'sender'
        ).order_by('created_at')

        # Filter by visibility for non-admin users
        if user.role != 'admin':
            messages_qs = [msg for msg in messages_qs if msg.can_view(user)]
        else:
            messages_qs = list(messages_qs)

        print(f"[get_chat_messages] Total messages to return: {len(messages_qs)}")

        # Debug: log media counts for messages that have media
        for msg in messages_qs:
            if msg.has_media:
                m2m = list(msg.media_files.all())
                fk  = list(msg.media_file.all())
                print(f"  → msg.id={msg.id} has_media=True | M2M={len(m2m)} FK={len(fk)}")

        # Update last read time
        if is_participant:
            ChatParticipant.objects.filter(
                chat_room=chat_room,
                user=user
            ).update(last_read_at=now())

        serializer = EnhancedMessageSerializer(
            messages_qs, many=True, context={'request': request}
        )

        return Response({
            'success': True,
            'messages': serializer.data,
            'total': len(messages_qs),
        })

    except Exception as e:
        return handle_exception(e, "Failed to get messages", request)
    
    
 

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_media(request, media_id):
    """Download a media file"""
    lang = get_user_language(request)
    try:
        user = request.user
        media_file = get_object_or_404(MediaFile, id=media_id)
        message = media_file.message
        
        # Check if user can view the message containing this media
        if not message.can_view(user) and user.role != 'admin':
            return Response({
                'error': ct("cannot_view_message", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check if user is participant in the chat room
        is_participant = message.chat_room.participants.filter(id=user.id).exists()
        if not is_participant and user.role != 'admin':
            return Response({
                'error': ct("not_participant", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Serve the file
        if media_file.file and os.path.exists(media_file.file.path):
            response = FileResponse(
                open(media_file.file.path, 'rb'),
                as_attachment=True,
                filename=media_file.file_name
            )
            return response
        else:
            raise Http404("File not found")
        
    except Http404:
        return Response({
            'error': ct("file_not_found", lang)
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return handle_exception(e, "Failed to download media", request)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stream_media(request, media_id):
    """Stream media file (for videos/audio in browser)"""
    lang = get_user_language(request)
    try:
        user = request.user
        media_file = get_object_or_404(MediaFile, id=media_id)
        message = media_file.message
        
        # Check permissions
        if not message.can_view(user) and user.role != 'admin':
            return Response({
                'error': ct("cannot_view_message", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        is_participant = message.chat_room.participants.filter(id=user.id).exists()
        if not is_participant and user.role != 'admin':
            return Response({
                'error': ct("not_participant", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Stream the file
        if media_file.file and os.path.exists(media_file.file.path):
            response = FileResponse(
                open(media_file.file.path, 'rb'),
                content_type=media_file.mime_type
            )
            response['Content-Disposition'] = f'inline; filename="{media_file.file_name}"'
            return response
        else:
            raise Http404("File not found")
        
    except Http404:
        return Response({
            'error': ct("file_not_found", lang)
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return handle_exception(e, "Failed to stream media", request)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_media(request, media_id):
    """Delete a media file (admin: permanent, others: hide for themselves)"""
    lang = get_user_language(request)
    try:
        user = request.user
        media_file = get_object_or_404(MediaFile, id=media_id)
        message = media_file.message
        
        # Check permissions
        if user.role == 'admin':
            # Admin can permanently delete
            if media_file.file and os.path.exists(media_file.file.path):
                os.remove(media_file.file.path)
            
            if media_file.thumbnail and os.path.exists(media_file.thumbnail.path):
                os.remove(media_file.thumbnail.path)
            
            media_file.delete()
            
            # Check if message still has media files
            if not message.media_files.exists():
                message.has_media = False
                message.save()
            
            return Response({
                'success': True,
                'message': ct("media_deleted_permanent", lang)
            })
        
        elif user == message.sender:
            # Sender can delete for everyone
            if media_file.file and os.path.exists(media_file.file.path):
                os.remove(media_file.file.path)
            
            if media_file.thumbnail and os.path.exists(media_file.thumbnail.path):
                os.remove(media_file.thumbnail.path)
            
            media_file.delete()
            
            # Check if message still has media files
            if not message.media_files.exists():
                message.has_media = False
                message.save()
            
            return Response({
                'success': True,
                'message': ct("media_deleted", lang)
            })
        
        else:
            # Other users can only hide for themselves
            MessageVisibilityOverride.objects.create(
                message=message,
                user=user,
                is_hidden=True
            )
            
            return Response({
                'success': True,
                'message': ct("media_hidden", lang)
            })
        
    except Exception as e:
        return handle_exception(e, "Failed to delete media", request)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_media_info(request, media_id):
    """Get detailed information about a media file"""
    lang = get_user_language(request)
    try:
        user = request.user
        media_file = get_object_or_404(MediaFile, id=media_id)
        message = media_file.message
        
        # Check if user can view the message
        if not message.can_view(user) and user.role != 'admin':
            return Response({
                'error': ct("cannot_view_message", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check if user is participant
        is_participant = message.chat_room.participants.filter(id=user.id).exists()
        if not is_participant and user.role != 'admin':
            return Response({
                'error': ct("not_participant", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = MediaFileSerializer(
            media_file, context={'request': request}
        )
        
        # Add additional info
        info = serializer.data
        info['message'] = {
            'id': message.id,
            'content': message.content,
            'created_at': message.created_at,
            'sender': {
                'id': message.sender.id,
                'name': message.sender.full_name,
                'role': message.sender.role
            }
        }
        info['chat_room'] = {
            'id': message.chat_room.id,
            'name': message.chat_room.name,
            'chat_type': message.chat_room.chat_type
        }
        
        return Response({
            'success': True,
            'media_info': info
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to get media info", request)
    
    

# ====== HELPER FUNCTIONS ======

def determine_file_type(filename, mime_type):
    """Determine the type of media file"""
    # Check by mime type first
    if mime_type.startswith('image/'):
        return MediaFileType.IMAGE
    elif mime_type.startswith('video/'):
        return MediaFileType.VIDEO
    elif mime_type.startswith('audio/'):
        # Check if it might be a voice note (you can add logic based on duration/size)
        return MediaFileType.AUDIO
    elif mime_type == 'application/pdf':
        return MediaFileType.PDF
    elif mime_type in [
        'application/msword',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    ]:
        return MediaFileType.WORD
    elif mime_type in [
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ]:
        return MediaFileType.EXCEL
    
    # Check by extension
    ext = os.path.splitext(filename)[1].lower()
    if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']:
        return MediaFileType.IMAGE
    elif ext in ['.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv', '.webm']:
        return MediaFileType.VIDEO
    elif ext in ['.mp3', '.wav', '.ogg', '.m4a', '.aac']:
        return MediaFileType.AUDIO
    elif ext in ['.pdf']:
        return MediaFileType.PDF
    elif ext in ['.doc', '.docx']:
        return MediaFileType.WORD
    elif ext in ['.xls', '.xlsx']:
        return MediaFileType.EXCEL
    
    return MediaFileType.DOCUMENT


def extract_metadata(media_file, file_obj):
    """Extract metadata based on file type"""
    try:
        if media_file.file_type == MediaFileType.IMAGE:
            # Extract image dimensions
            img = Image.open(file_obj)
            media_file.dimensions = f"{img.width}x{img.height}"
            
            # Create thumbnail
            img.thumbnail((200, 200))
            thumb_path = f"thumbnails/{media_file.file.name.split('/')[-1]}"
            thumb_full_path = os.path.join(settings.MEDIA_ROOT, thumb_path)
            os.makedirs(os.path.dirname(thumb_full_path), exist_ok=True)
            img.save(thumb_full_path)
            media_file.thumbnail = thumb_path
            
        elif media_file.file_type == MediaFileType.VIDEO:
            # Extract video duration and dimensions
            clip = VideoFileClip(file_obj.temporary_file_path())
            media_file.duration = clip.duration
            media_file.dimensions = f"{clip.size[0]}x{clip.size[1]}"
            clip.close()
            
        elif media_file.file_type == MediaFileType.AUDIO:
            # Extract audio duration
            audio = AudioSegment.from_file(file_obj)
            media_file.duration = len(audio) / 1000  # Convert to seconds
            
        elif media_file.file_type == MediaFileType.PDF:
            # Extract page count
            pdf_reader = PyPDF2.PdfReader(file_obj)
            media_file.page_count = len(pdf_reader.pages)
            
        elif media_file.file_type == MediaFileType.WORD:
            # Extract word count (approximate)
            doc = Document(file_obj)
            word_count = 0
            for paragraph in doc.paragraphs:
                word_count += len(paragraph.text.split())
            media_file.word_count = word_count
            
        elif media_file.file_type == MediaFileType.EXCEL:
            # Extract sheet count
            wb = load_workbook(file_obj, read_only=True)
            media_file.page_count = len(wb.sheetnames)
        
        media_file.save()
        
    except Exception as e:
        logger.error(f"Error extracting metadata: {e}")
        
        

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_get_chat_by_id(request, room_id):
    """Admin: Get a specific chat by ID"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        if user.role != 'admin':
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id)
        
        serializer = ChatRoomSerializer(chat_room, context={'request': request})
        data = serializer.data
        data['participants'] = ChatParticipantSerializer(
            chat_room.chat_participants.select_related('user'), 
            many=True, 
            context={'request': request}
        ).data
        data['message_count'] = chat_room.messages.filter(is_deleted=False).count()
        
        return Response({
            'success': True,
            'chat': data
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to get chat", request)
    
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_get_chat_by_id(request, room_id):
    """User: Get a specific chat by ID"""
    lang = get_user_language(request)
    try:
        user = request.user
        
        if user.role != 'admin' and not ChatRoom.objects.filter(id=room_id, participants=user).exists():
            return Response({
                'error': ct("admin_required", lang)
            }, status=status.HTTP_403_FORBIDDEN)
        
        chat_room = get_object_or_404(ChatRoom, id=room_id)
        
        serializer = ChatRoomSerializer(chat_room, context={'request': request})
        data = serializer.data
        data['participants'] = ChatParticipantSerializer(
            chat_room.chat_participants.select_related('user'), 
            many=True, 
            context={'request': request}
        ).data
        data['message_count'] = chat_room.messages.filter(is_deleted=False).count()
        
        return Response({
            'success': True,
            'chat': data
        })
        
    except Exception as e:
        return handle_exception(e, "Failed to get chat", request)